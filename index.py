import pandas as pd
import json
import requests
import openpyxl
import plotly.express as px
#cidade=input('Diga a cidade que deseja saber a temperatura (Say the city you wish to know the temperature): ')
key='9a40fecccf7d49d1935151140240211'
request=json.loads(requests.get(f'http://api.weatherapi.com/v1/current.json?key={key}&q=Santos&aqi=no').content.decode('utf-8'))
time=request['location']['localtime']
temp=float(request['current']['temp_c'])
print(f"Temperatura:{temp}°C\nData e hora:{time}")
try:
    wb=openpyxl.load_workbook('Test.xlsx')
    ws=wb.active
    ws.append([temp,time])
except:
    wb=openpyxl.Workbook()
    ws=wb.active
    ws['A1']='temperatura'
    ws['A2']=temp
    ws['B1']='Horário'
    ws['B2']=time
wb.save('Test.xlsx')
arquivo=pd.read_excel('Test.xlsx')
grafico=px.bar(arquivo,x='Horário', y='temperatura')
grafico.show()
